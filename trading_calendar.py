# ================================ IMPORTS ================================

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import sys
import json
import shutil
import threading
import time
import tempfile
from datetime import datetime, timedelta
from collections import defaultdict
import calendar

# ================================ DEPENDENCY MANAGEMENT ================================

def check_dependencies():
    required_packages = ['openpyxl', 'requests']
    missing_packages = []
    
    for package in required_packages:
        try:
            __import__(package)
        except ImportError:
            missing_packages.append(package)
    
    if missing_packages:
        print(f"Installing required packages: {', '.join(missing_packages)}")
        import subprocess
        for package in missing_packages:
            try:
                subprocess.check_call([sys.executable, '-m', 'pip', 'install', package])
                print(f"{package} installed successfully")
            except subprocess.CalledProcessError as e:
                print(f"Failed to install {package}: {e}")
                return False
    
    return True

if not check_dependencies():
    print("Failed to install required dependencies")
    input("Press Enter to exit...")
    sys.exit(1)

import openpyxl
import requests

# ================================ CONFIGURATION ================================

DEFAULT_CONFIG = {
    "theme": {
        "bg_dark": "#0f0f0f",
        "bg_card": "#1a1a1a",
        "bg_accent": "#262626",
        "text_primary": "#ffffff",
        "text_secondary": "#9ca3af",
        "text_muted": "#6b7280",
        "accent_blue": "#3b82f6",
        "accent_green": "#10b981",
        "accent_red": "#ef4444",
        "accent_yellow": "#f59e0b",
        "border": "#374151",
        "hover": "#1f2937"
    },
    "excel_columns": {
        "datetime": "Date/Time",
        "pnl": "P&L USD",
        "trade_num": "Trade #",
        "type": "Type"
    },
    "sheet_names": [
        "List of trades",
        "list of trades"
    ],
    "ui": {
        "window_width": 1200,
        "window_height": 800,
        "loading_timeout": 2000,
        "progress_update_interval": 10
    },
    "version": "1.0.0"
}

# ================================ DATA MANAGEMENT ================================

class DataManager:
    
    def __init__(self, data_dir):
        self.data_dir = data_dir
        self.ensure_data_directory()
    
    def ensure_data_directory(self):
        if not os.path.exists(self.data_dir):
            os.makedirs(self.data_dir, exist_ok=True)
    
    def save_trading_data(self, filename, trades, daily_pnl, stats):
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_name = os.path.splitext(os.path.basename(filename))[0]
            save_file = os.path.join(self.data_dir, f"{base_name}_{timestamp}.json")
            
            serializable_trades = []
            for trade in trades:
                trade_data = trade.copy()
                if 'date' in trade_data:
                    trade_data['date'] = trade_data['date'].isoformat()
                serializable_trades.append(trade_data)
            
            data = {
                'timestamp': timestamp,
                'original_filename': os.path.basename(filename),
                'trades': serializable_trades,
                'daily_pnl': dict(daily_pnl),
                'stats': stats,
                'total_trades': len(trades)
            }
            
            with open(save_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2)
            
            print(f"Trading data saved: {os.path.basename(save_file)}")
            return save_file
            
        except (IOError, OSError, json.JSONEncodeError) as e:
            print(f"Error saving data: {e}")
            return None
    
    def get_data_history(self):
        try:
            if not os.path.exists(self.data_dir):
                return []
            
            files = []
            for filename in os.listdir(self.data_dir):
                if filename.endswith('.json'):
                    filepath = os.path.join(self.data_dir, filename)
                    files.append({
                        'filename': filename,
                        'path': filepath,
                        'modified': os.path.getmtime(filepath)
                    })
            
            files.sort(key=lambda x: x['modified'], reverse=True)
            return files
            
        except (IOError, OSError) as e:
            print(f"Error getting data history: {e}")
            return []

# ================================ TRADE PROCESSING ================================

class TradeProcessor:
    
    def __init__(self, config=None):
        self.config = config or DEFAULT_CONFIG
        self.trades = []
        self.daily_pnl = defaultdict(float)
        self.stats = {}
    
    def process_file(self, file_path, progress_callback=None):
        try:
            # Quick file size check to optimize loading strategy
            file_size = os.path.getsize(file_path)
            is_small_file = file_size < 1024 * 1024  # Less than 1MB
            
            if progress_callback and not is_small_file:
                progress_callback("Opening Excel file...")
            
            # Use read_only mode for better performance
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True, keep_links=False)
            
            if progress_callback and not is_small_file:
                progress_callback("Scanning for trading data sheets...")
            
            sheet = self._find_trades_sheet(wb, progress_callback if not is_small_file else None)
            if not sheet:
                raise ValueError("Could not find trading data sheet")
            
            if progress_callback and not is_small_file:
                progress_callback("Analyzing column headers...")
            
            headers = self._analyze_headers(sheet, progress_callback if not is_small_file else None)
            
            # For small files, skip intermediate progress updates
            self._process_trades(sheet, headers, progress_callback if not is_small_file else None)
            
            self.calculate_stats()
            
            # Close workbook to free memory
            wb.close()
            
            return {
                'trades': self.trades,
                'daily_pnl': dict(self.daily_pnl),
                'stats': self.stats
            }
            
        except Exception as e:
            raise RuntimeError(f"Error processing Excel file: {str(e)}") from e
    
    def _find_trades_sheet(self, wb, progress_callback=None):
        sheet_names = self.config.get("sheet_names", ["List of trades"])
        
        for sheet_name in wb.sheetnames:
            for target_name in sheet_names:
                if target_name.lower() in sheet_name.lower():
                    if progress_callback:
                        progress_callback(f"Found sheet: '{sheet_name}'")
                    return wb[sheet_name]
        
        return None
    
    def _analyze_headers(self, sheet, progress_callback=None):
        headers = {}
        column_config = self.config.get("excel_columns", {})
        first_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        
        for i, header in enumerate(first_row):
            if header:
                header_str = str(header).strip()
                
                for key, expected_name in column_config.items():
                    if header_str == expected_name:
                        headers[key] = i
                        if progress_callback:
                            progress_callback(f"Found {expected_name} column")
                        break
        
        required_columns = ['datetime', 'pnl']
        missing_columns = [col for col in required_columns if col not in headers]
        
        if missing_columns:
            column_names = [column_config.get(col, col) for col in missing_columns]
            raise ValueError(f"Missing required columns: {', '.join(column_names)}")
        
        return headers
    
    def _process_trades(self, sheet, headers, progress_callback=None):
        self.trades = []
        self.daily_pnl = defaultdict(float)
        processed_trade_numbers = set()
        
        # Read all rows into memory at once for faster processing
        if progress_callback:
            progress_callback("Reading Excel data into memory...")
        
        all_rows = list(sheet.iter_rows(min_row=2, values_only=True))
        total_rows = len(all_rows)
        
        if total_rows == 0:
            if progress_callback:
                progress_callback("No data rows found")
            return
        
        # Pre-allocate lists for better performance
        trades_list = []
        daily_pnl_dict = defaultdict(float)
        
        # Smart progress reporting - less overhead for small files
        if total_rows < 1000:
            # For small files, minimal progress updates
            progress_interval = max(1, total_rows // 10)  # Max 10 updates
            show_progress = total_rows > 100  # Only show progress if >100 rows
        else:
            # For large files, more frequent updates
            progress_interval = max(1, total_rows // 100)  # Max 100 updates
            show_progress = True
        
        # Get column indices once
        datetime_col = headers['datetime']
        pnl_col = headers['pnl']
        trade_num_col = headers.get('trade_num')
        
        if progress_callback and show_progress:
            progress_callback("Processing trade data...")
        
        for i, row in enumerate(all_rows):
            if progress_callback and show_progress and i % progress_interval == 0:
                progress = int((i / total_rows) * 100)
                progress_callback(f"Processing trades... {progress}% ({i}/{total_rows})")
            
            try:
                # Fast inline processing instead of method calls
                trade_num = row[trade_num_col] if trade_num_col is not None else None
                datetime_val = row[datetime_col]
                pnl_val = row[pnl_col]
                
                # Skip invalid rows quickly
                if not datetime_val or pnl_val is None:
                    continue
                
                # Skip duplicate trade numbers
                if trade_num and trade_num in processed_trade_numbers:
                    continue
                
                # Fast date parsing
                if isinstance(datetime_val, datetime):
                    trade_date = datetime_val.date()
                else:
                    try:
                        date_str = str(datetime_val).split()[0]
                        trade_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                    except (ValueError, TypeError):
                        continue
                
                # Validate date range quickly
                if trade_date > datetime.now().date() or trade_date.year < 2000:
                    continue
                
                # Fast PnL parsing
                if isinstance(pnl_val, (int, float)):
                    pnl = float(pnl_val)
                elif isinstance(pnl_val, str):
                    cleaned = pnl_val.replace(',', '').replace('$', '').strip()
                    if not cleaned or cleaned == '-':
                        continue
                    try:
                        pnl = float(cleaned)
                    except ValueError:
                        continue
                else:
                    continue
                
                # Add trade number to processed set
                if trade_num:
                    processed_trade_numbers.add(trade_num)
                
                # Create trade record
                trade = {
                    'date': trade_date,
                    'pnl': pnl,
                    'trade_num': trade_num
                }
                
                trades_list.append(trade)
                date_key = trade_date.strftime('%Y-%m-%d')
                daily_pnl_dict[date_key] += pnl
                
            except (ValueError, TypeError, AttributeError):
                continue
        
        # Assign results
        self.trades = trades_list
        self.daily_pnl = daily_pnl_dict
        
        if progress_callback:
            progress_callback(f"Successfully processed {len(self.trades)} unique trades")
    
    def _process_single_trade(self, row, headers, processed_trade_numbers):
        trade_num = row[headers['trade_num']] if 'trade_num' in headers else None
        datetime_val = row[headers['datetime']]
        pnl_val = row[headers['pnl']]
        
        if not datetime_val or pnl_val is None:
            return None
        
        if trade_num and trade_num in processed_trade_numbers:
            return None
        
        trade_date = self._parse_date(datetime_val)
        if not trade_date:
            return None
        
        pnl = self._parse_pnl(pnl_val)
        if pnl is None:
            return None
        
        if trade_num:
            processed_trade_numbers.add(trade_num)
        
        return {
            'date': trade_date,
            'pnl': pnl,
            'trade_num': trade_num
        }
    
    def _parse_date(self, datetime_val):
        try:
            if isinstance(datetime_val, datetime):
                parsed_date = datetime_val.date()
            else:
                date_str = str(datetime_val).split()[0]
                parsed_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            
            if parsed_date > datetime.now().date():
                return None
            if parsed_date.year < 2000:
                return None
                
            return parsed_date
        except (ValueError, TypeError, AttributeError):
            return None
    
    def _parse_pnl(self, pnl_val):
        try:
            if isinstance(pnl_val, (int, float)):
                return float(pnl_val)
            elif isinstance(pnl_val, str):
                cleaned = pnl_val.replace(',', '').replace('$', '').strip()
                if not cleaned or cleaned == '-':
                    return None
                return float(cleaned)
            else:
                return None
        except (ValueError, TypeError):
            return None
    
    def calculate_stats(self):
        if not self.trades:
            self.stats = {
                'total_pnl': 0.0,
                'win_rate': 0.0,
                'avg_daily': 0.0,
                'total_trades': 0
            }
            return
        
        # Fast statistics calculation
        total_trades = len(self.trades)
        total_pnl = 0.0
        winning_trades = 0
        
        # Single pass through trades for efficiency
        for trade in self.trades:
            pnl = trade['pnl']
            total_pnl += pnl
            if pnl > 0:
                winning_trades += 1
        
        win_rate = (winning_trades / total_trades * 100) if total_trades > 0 else 0
        
        # Count trading days efficiently
        trading_days = sum(1 for pnl in self.daily_pnl.values() if pnl != 0)
        avg_daily = total_pnl / trading_days if trading_days > 0 else 0
        
        self.stats = {
            'total_pnl': total_pnl,
            'win_rate': win_rate,
            'avg_daily': avg_daily,
            'total_trades': total_trades
        }

# ================================ MAIN APPLICATION ================================

class StrategyAnalyzer:
    
    def __init__(self, root):
        self.root = root
        self.root.title("Strategy Analyzer")
        
        self.check_launcher_updates()
        
        self.config = DEFAULT_CONFIG
        ui_config = self.config.get("ui", {})
        
        self.root.geometry(f"{ui_config.get('window_width', 1200)}x{ui_config.get('window_height', 800)}")
        self.root.configure(bg=self.config['theme']['bg_dark'])
        self.root.resizable(True, True)
        
        self.CURRENT_VERSION = os.environ.get('STRATEGY_ANALYZER_VERSION', self.config.get('version', '1.0.0'))
        
        self.validate_version()
        
        # Try to find the correct data directory
        data_dir = os.environ.get('STRATEGY_ANALYZER_DATA_DIR')
        if not data_dir:
            # If not set via launcher, use the standard AppData location
            appdata_dir = os.path.join(os.path.expanduser("~"), "AppData", "Roaming", "StrategyAnalyzer", "data")
            if os.path.exists(appdata_dir):
                data_dir = appdata_dir
            else:
                # Fallback to local data folder
                data_dir = os.path.join(os.getcwd(), 'data')
        
        self.data_manager = DataManager(data_dir)
        
        self.trade_processor = TradeProcessor(self.config)
        
        self.theme = self.config['theme']
        
        self.trades = []
        self.daily_pnl = defaultdict(float)
        self.stats = {}
        self.current_month = datetime.now().month
        self.current_year = datetime.now().year
        self.current_file = None
        self.trades_by_date = {}
        self.performance_metrics = {}
        self.show_performance_logs = False
        
        self.setup_ui()
    
    def check_launcher_updates(self):
        if os.environ.get('STRATEGY_ANALYZER_VERSION') == 'DEV':
            print("Development mode - skipping launcher updates")
            return
            
        try:
            import tempfile
            import subprocess
            
            launcher_path = None
            
            if len(sys.argv) > 0:
                potential_paths = [
                    os.path.join(os.path.expanduser("~"), "Desktop", "setup_backtestcalendar.bat"),
                    os.path.join(os.path.expanduser("~"), "Downloads", "setup_backtestcalendar.bat"),
                    os.path.join(os.getcwd(), "setup_backtestcalendar.bat"),
                    os.path.join(os.path.dirname(sys.argv[0]), "setup_backtestcalendar.bat")
                ]
                
                for path in potential_paths:
                    if os.path.exists(path):
                        launcher_path = path
                        break
            
            if not launcher_path:
                return
            
            print(f"Checking for launcher updates: {launcher_path}")
            
            launcher_url = "https://raw.githubusercontent.com/Stellarr-r/trading-calendar-app/main/setup_backtestcalendar.bat"
            temp_launcher = os.path.join(tempfile.gettempdir(), "setup_backtestcalendar_new.bat")
            
            try:
                import requests
                response = requests.get(launcher_url, timeout=10)
                response.raise_for_status()
                
                with open(temp_launcher, 'w', encoding='utf-8') as f:
                    f.write(response.text)
                
                if os.path.exists(temp_launcher):
                    needs_update = False
                    
                    if os.path.exists(launcher_path):
                        with open(launcher_path, 'rb') as f1, open(temp_launcher, 'rb') as f2:
                            if f1.read() != f2.read():
                                needs_update = True
                    else:
                        needs_update = True
                    
                    if needs_update:
                        print("Launcher update available - updating...")
                        
                        if launcher_path and os.path.dirname(launcher_path):
                            try:
                                shutil.copy2(temp_launcher, launcher_path)
                                print(f"Launcher updated successfully: {launcher_path}")
                            except Exception as e:
                                print(f"Could not update original launcher: {e}")
                        
                        try:
                            script_dir = os.path.dirname(os.path.abspath(__file__))
                            backup_launcher = os.path.join(script_dir, "setup_backtestcalendar.bat")
                            shutil.copy2(temp_launcher, backup_launcher)
                            print(f"Backup launcher created: {backup_launcher}")
                        except Exception as e:
                            print(f"Could not create backup launcher: {e}")
                            
                    else:
                        print("Launcher is up to date")
                        
                        try:
                            script_dir = os.path.dirname(os.path.abspath(__file__))
                            backup_launcher = os.path.join(script_dir, "setup_backtestcalendar.bat")
                            if not os.path.exists(backup_launcher):
                                shutil.copy2(temp_launcher, backup_launcher)
                                print(f"Backup launcher created: {backup_launcher}")
                        except Exception as e:
                            print(f"Could not create backup launcher: {e}")
                    
                    os.remove(temp_launcher)
                    
            except Exception as e:
                print(f"Could not update launcher: {e}")
                if os.path.exists(temp_launcher):
                    os.remove(temp_launcher)
                
        except Exception as e:
            print(f"Launcher update check failed: {e}")
    
    def validate_version(self):
        import re
        
        version_pattern = r'^(\d+\.\d+\.\d+|DEV)$'
        if not re.match(version_pattern, self.CURRENT_VERSION):
            print(f"Warning: Invalid version format '{self.CURRENT_VERSION}'. Using fallback.")
            self.CURRENT_VERSION = "1.0.0"
        
        print(f"Strategy Analyzer v{self.CURRENT_VERSION} starting...")
        
        version_source = os.environ.get('STRATEGY_ANALYZER_VERSION')
        if version_source:
            print(f"Version loaded from Strategy Analyzer launcher: {version_source}")
        else:
            print(f"Using default version (run via Strategy Analyzer launcher for version management)")
        
    def setup_ui(self):
        main_container = tk.Frame(self.root, bg=self.theme['bg_dark'])
        main_container.pack(fill='both', expand=True, padx=0, pady=0)
        
        self.create_top_bar(main_container)
        
        content_frame = tk.Frame(main_container, bg=self.theme['bg_dark'])
        content_frame.pack(fill='both', expand=True, padx=20, pady=0)
        
        loading_label = tk.Label(content_frame, 
                               text="Loading...", 
                               font=('Inter', 12),
                               fg=self.theme['text_muted'],
                               bg=self.theme['bg_dark'])
        loading_label.pack(pady=50)
        
        self.create_heavy_ui_elements(content_frame, loading_label)
        
    def create_heavy_ui_elements(self, content_frame, loading_label):
        loading_label.destroy()
        
        self.create_stats_section(content_frame)
        
        self.create_calendar_section(content_frame)
        
        self.create_bottom_info(content_frame)
        
    def create_top_bar(self, parent):
        top_bar = tk.Frame(parent, bg=self.theme['bg_card'], height=70)
        top_bar.pack(fill='x', padx=0, pady=0)
        top_bar.pack_propagate(False)
        
        left_frame = tk.Frame(top_bar, bg=self.theme['bg_card'])
        left_frame.pack(side='left', fill='y', padx=20)
        
        title = tk.Label(left_frame,
                        text="Strategy Analyzer",
                        font=('Inter', 20, 'bold'),
                        fg=self.theme['text_primary'],
                        bg=self.theme['bg_card'])
        title.pack(side='left', pady=20)
        
        version_label = tk.Label(left_frame,
                               text=f"v{self.CURRENT_VERSION}",
                               font=('Inter', 10),
                               fg=self.theme['text_muted'],
                               bg=self.theme['bg_card'])
        version_label.pack(side='left', padx=(10, 0), pady=20)
        
        right_frame = tk.Frame(top_bar, bg=self.theme['bg_card'])
        right_frame.pack(side='right', fill='y', padx=20)
        
        settings_btn = tk.Button(right_frame,
                                text="⚙",
                                font=('Inter', 12),
                                bg=self.theme['bg_accent'],
                                fg=self.theme['text_primary'],
                                border=0,
                                width=3,
                                pady=8,
                                cursor='hand2',
                                activebackground=self.theme['hover'],
                                command=self.show_settings)
        settings_btn.pack(side='right', pady=20, padx=(0, 10))
        
        load_btn = tk.Button(right_frame,
                            text="Load File",
                            font=('Inter', 10, 'bold'),
                            bg=self.theme['accent_blue'],
                            fg='white',
                            border=0,
                            padx=20,
                            pady=8,
                            cursor='hand2',
                            activebackground='#2563eb',
                            command=self.load_file)
        load_btn.pack(side='right', pady=20)
        
        self.file_status = tk.Label(left_frame,
                                   text="No file loaded",
                                   font=('Inter', 9),
                                   fg=self.theme['text_muted'],
                                   bg=self.theme['bg_card'])
        self.file_status.pack(side='left', padx=(20, 0), pady=20)
        
    def create_stats_section(self, parent):
        stats_container = tk.Frame(parent, bg=self.theme['bg_dark'])
        stats_container.pack(fill='x', pady=20)
        
        stats_grid = tk.Frame(stats_container, bg=self.theme['bg_dark'])
        stats_grid.pack(fill='x')
        
        self.stat_cards = {}
        stats_config = [
            ("Total P&L", "$0.00", "total_pnl"),
            ("Win Rate", "0%", "win_rate"),
            ("Avg Daily", "$0.00", "avg_daily"),
            ("Total Trades", "0", "total_trades")
        ]
        
        for i, (label, value, key) in enumerate(stats_config):
            card = self.create_stat_card(stats_grid, label, value)
            card.grid(row=0, column=i, padx=(0, 15 if i < 3 else 0), sticky='ew')
            self.stat_cards[key] = card
            stats_grid.grid_columnconfigure(i, weight=1)
    
    def create_stat_card(self, parent, label, value):
        card = tk.Frame(parent, 
                       bg=self.theme['bg_card'],
                       relief='flat',
                       bd=0)
        
        content = tk.Frame(card, bg=self.theme['bg_card'])
        content.pack(fill='both', expand=True, padx=20, pady=20)
        
        value_label = tk.Label(content,
                              text=value,
                              font=('Inter', 24, 'bold'),
                              fg=self.theme['text_primary'],
                              bg=self.theme['bg_card'])
        value_label.pack(anchor='w')
        
        label_widget = tk.Label(content,
                               text=label,
                               font=('Inter', 11),
                               fg=self.theme['text_secondary'],
                               bg=self.theme['bg_card'])
        label_widget.pack(anchor='w', pady=(5, 0))
        
        card.value_label = value_label
        card.label_widget = label_widget
        
        return card
    
    def create_calendar_section(self, parent):
        calendar_container = tk.Frame(parent, bg=self.theme['bg_dark'])
        calendar_container.pack(fill='both', expand=True, pady=(20, 0))
        
        cal_header = tk.Frame(calendar_container, bg=self.theme['bg_dark'])
        cal_header.pack(fill='x', pady=(0, 20))
        
        nav_frame = tk.Frame(cal_header, bg=self.theme['bg_dark'])
        nav_frame.pack()
        
        prev_btn = tk.Button(nav_frame,
                            text="‹",
                            font=('Inter', 18),
                            bg=self.theme['bg_accent'],
                            fg=self.theme['text_primary'],
                            border=0,
                            width=3,
                            pady=8,
                            cursor='hand2',
                            activebackground=self.theme['hover'],
                            command=self.prev_month)
        prev_btn.pack(side='left')
        
        month_name = calendar.month_name[self.current_month]
        self.month_display = tk.Label(nav_frame,
                                     text=f"{month_name} {self.current_year}",
                                     font=('Inter', 16, 'bold'),
                                     fg=self.theme['text_primary'],
                                     bg=self.theme['bg_dark'],
                                     padx=30)
        self.month_display.pack(side='left')
        
        next_btn = tk.Button(nav_frame,
                            text="›",
                            font=('Inter', 18),
                            bg=self.theme['bg_accent'],
                            fg=self.theme['text_primary'],
                            border=0,
                            width=3,
                            pady=8,
                            cursor='hand2',
                            activebackground=self.theme['hover'],
                            command=self.next_month)
        next_btn.pack(side='left')
        
        self.calendar_frame = tk.Frame(calendar_container, bg=self.theme['bg_dark'])
        self.calendar_frame.pack(fill='both', expand=True)
        
        self.setup_calendar_grid()
        
    def setup_calendar_grid(self):
        days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
        for i, day in enumerate(days):
            header = tk.Label(self.calendar_frame,
                             text=day,
                             font=('Inter', 11, 'bold'),
                             fg=self.theme['text_secondary'],
                             bg=self.theme['bg_accent'],
                             pady=12)
            header.grid(row=0, column=i, sticky='ew', padx=1, pady=1)
        
        for i in range(7):
            self.calendar_frame.grid_columnconfigure(i, weight=1)
        for i in range(7):
            self.calendar_frame.grid_rowconfigure(i, weight=1)
            
        self.update_calendar()
    
    def create_day_cell(self, day, pnl=0, row=0, col=0):
        date_key = f"{self.current_year}-{self.current_month:02d}-{day:02d}"
        day_trades = self.trades_by_date.get(date_key, [])
        trade_count = len(day_trades)
        
        if pnl > 0:
            bg_color = self.theme['accent_green']
            text_color = 'white'
            hover_color = '#059669'
        elif pnl < 0:
            bg_color = self.theme['accent_red']
            text_color = 'white'
            hover_color = '#dc2626'
        else:
            bg_color = self.theme['bg_card']
            text_color = self.theme['text_secondary']
            hover_color = self.theme['hover']
        
        cell = tk.Frame(self.calendar_frame,
                       bg=bg_color,
                       relief='flat',
                       bd=0,
                       cursor='hand2' if trade_count > 0 else 'arrow')
        cell.grid(row=row, column=col, sticky='nsew', padx=1, pady=1)
        
        content = tk.Frame(cell, bg=bg_color)
        content.pack(fill='both', expand=True, padx=8, pady=8)
        
        day_label = tk.Label(content,
                            text=str(day),
                            font=('Inter', 12, 'bold'),
                            fg=text_color,
                            bg=bg_color)
        day_label.pack(anchor='nw')
        
        if trade_count > 0:
            pnl_text = f"${pnl:,.0f}" if abs(pnl) >= 1 else f"${pnl:.2f}"
            pnl_label = tk.Label(content,
                                text=pnl_text,
                                font=('Inter', 9, 'bold'),
                                fg=text_color,
                                bg=bg_color)
            pnl_label.pack(anchor='center')
            
            trades_label = tk.Label(content,
                                   text=f"{trade_count} trade{'s' if trade_count != 1 else ''}",
                                   font=('Inter', 8),
                                   fg=text_color,
                                   bg=bg_color)
            trades_label.pack(anchor='sw')
            
            widgets_list = [cell, content, day_label, pnl_label, trades_label]
        else:
            widgets_list = [cell, content, day_label]
        
        # FIXED: Create event handlers with proper closure capture using default arguments
        def create_click_handler(dk, dt):
            def handle_click(e):
                if self.show_performance_logs:
                    print(f"Cell clicked: {dk}, trades: {len(dt) if dt else 0}")
                if dt:
                    self.show_day_details(dk, dt)
            return handle_click
        
        def create_hover_handlers(widgets, orig_bg, hover_bg):
            def on_enter(e):
                for widget in widgets:
                    widget.config(bg=hover_bg)
            def on_leave(e):
                for widget in widgets:
                    widget.config(bg=orig_bg)
            return on_enter, on_leave
        
        # FIXED: Always bind click event to ALL widgets so clicking anywhere works
        click_handler = create_click_handler(date_key, day_trades)
        for widget in widgets_list:
            widget.bind('<Button-1>', click_handler)
        
        # Only add hover effects for cells with trades
        if trade_count > 0:
            on_enter, on_leave = create_hover_handlers(widgets_list, bg_color, hover_color)
            for widget in widgets_list:
                widget.bind('<Enter>', on_enter)
                widget.bind('<Leave>', on_leave)
    
    def create_optimized_day_cell(self, day, pnl, day_trades, row, col):
        """Optimized version that accepts pre-calculated data to avoid lookups"""
        trade_count = len(day_trades)
        
        # OPTIMIZED: Use pre-calculated values instead of recalculating
        if pnl > 0:
            bg_color = self.theme['accent_green']
            text_color = 'white'
            hover_color = '#059669'
        elif pnl < 0:
            bg_color = self.theme['accent_red']
            text_color = 'white'
            hover_color = '#dc2626'
        else:
            bg_color = self.theme['bg_card']
            text_color = self.theme['text_secondary']
            hover_color = self.theme['hover']
        
        cell = tk.Frame(self.calendar_frame,
                       bg=bg_color,
                       relief='flat',
                       bd=0,
                       cursor='hand2' if trade_count > 0 else 'arrow')
        cell.grid(row=row, column=col, sticky='nsew', padx=1, pady=1)
        
        content = tk.Frame(cell, bg=bg_color)
        content.pack(fill='both', expand=True, padx=8, pady=8)
        
        day_label = tk.Label(content,
                            text=str(day),
                            font=('Inter', 12, 'bold'),
                            fg=text_color,
                            bg=bg_color)
        day_label.pack(anchor='nw')
        
        # Pre-calculate date_key for all cells (needed for click handler)
        date_key = f"{self.current_year}-{self.current_month:02d}-{day:02d}"
        
        if trade_count > 0:
            pnl_text = f"${pnl:,.0f}" if abs(pnl) >= 1 else f"${pnl:.2f}"
            pnl_label = tk.Label(content,
                                text=pnl_text,
                                font=('Inter', 9, 'bold'),
                                fg=text_color,
                                bg=bg_color)
            pnl_label.pack(anchor='center')
            
            trades_label = tk.Label(content,
                                   text=f"{trade_count} trade{'s' if trade_count != 1 else ''}",
                                   font=('Inter', 8),
                                   fg=text_color,
                                   bg=bg_color)
            trades_label.pack(anchor='sw')
            
            widgets_list = [cell, content, day_label, pnl_label, trades_label]
        else:
            widgets_list = [cell, content, day_label]
        
        # FIXED: Create event handlers with proper closure capture using default arguments
        def create_click_handler(dk, dt):
            def handle_click(e):
                if self.show_performance_logs:
                    print(f"Cell clicked: {dk}, trades: {len(dt) if dt else 0}")
                if dt:
                    self.show_day_details(dk, dt)
            return handle_click
        
        def create_hover_handlers(widgets, orig_bg, hover_bg):
            def on_enter(e):
                for widget in widgets:
                    widget.config(bg=hover_bg)
            def on_leave(e):
                for widget in widgets:
                    widget.config(bg=orig_bg)
            return on_enter, on_leave
        
        # FIXED: Always bind click event to ALL widgets so clicking anywhere works
        click_handler = create_click_handler(date_key, day_trades)
        for widget in widgets_list:
            widget.bind('<Button-1>', click_handler)
        
        # Only add hover effects for cells with trades
        if trade_count > 0:
            on_enter, on_leave = create_hover_handlers(widgets_list, bg_color, hover_color)
            for widget in widgets_list:
                widget.bind('<Enter>', on_enter)
                widget.bind('<Leave>', on_leave)
    
    def show_settings(self):
        settings_window = tk.Toplevel(self.root)
        settings_window.title("Settings")
        settings_window.geometry("500x450")
        settings_window.configure(bg=self.theme['bg_dark'])
        settings_window.transient(self.root)
        settings_window.grab_set()
        
        settings_window.geometry("+{}+{}".format(
            int(self.root.winfo_screenwidth()/2 - 250),
            int(self.root.winfo_screenheight()/2 - 225)
        ))
        
        main_frame = tk.Frame(settings_window, bg=self.theme['bg_dark'])
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        title_label = tk.Label(main_frame,
                              text="Settings",
                              font=('Inter', 18, 'bold'),
                              fg=self.theme['text_primary'],
                              bg=self.theme['bg_dark'])
        title_label.pack(anchor='w', pady=(0, 20))
        
        # Performance section
        perf_frame = tk.Frame(main_frame, bg=self.theme['bg_card'])
        perf_frame.pack(fill='x', pady=(0, 15))
        
        perf_content = tk.Frame(perf_frame, bg=self.theme['bg_card'])
        perf_content.pack(fill='x', padx=15, pady=15)
        
        tk.Label(perf_content,
                text="Performance",
                font=('Inter', 14, 'bold'),
                fg=self.theme['text_primary'],
                bg=self.theme['bg_card']).pack(anchor='w', pady=(0, 10))
        
        # Performance logging checkbox
        perf_log_frame = tk.Frame(perf_content, bg=self.theme['bg_card'])
        perf_log_frame.pack(fill='x', pady=5)
        
        self.perf_log_var = tk.BooleanVar(value=self.show_performance_logs)
        perf_checkbox = tk.Checkbutton(perf_log_frame,
                                      text="Show performance logs in console",
                                      variable=self.perf_log_var,
                                      font=('Inter', 11),
                                      fg=self.theme['text_primary'],
                                      bg=self.theme['bg_card'],
                                      selectcolor=self.theme['bg_accent'],
                                      activebackground=self.theme['bg_card'],
                                      activeforeground=self.theme['text_primary'],
                                      command=self.toggle_performance_logs)
        perf_checkbox.pack(anchor='w')
        
        tk.Label(perf_content,
                text="Enable detailed timing information for calendar updates",
                font=('Inter', 9),
                fg=self.theme['text_muted'],
                bg=self.theme['bg_card']).pack(anchor='w', padx=(25, 0))
        
        # Data Management section
        data_frame = tk.Frame(main_frame, bg=self.theme['bg_card'])
        data_frame.pack(fill='x', pady=(0, 15))
        
        data_content = tk.Frame(data_frame, bg=self.theme['bg_card'])
        data_content.pack(fill='x', padx=15, pady=15)
        
        tk.Label(data_content,
                text="Data Management",
                font=('Inter', 14, 'bold'),
                fg=self.theme['text_primary'],
                bg=self.theme['bg_card']).pack(anchor='w', pady=(0, 10))
        
        # Data folder information
        folder_size = self.get_data_folder_size()
        file_count = self.get_data_file_count()
        formatted_size = self.format_file_size(folder_size)
        
        info_frame = tk.Frame(data_content, bg=self.theme['bg_card'])
        info_frame.pack(fill='x', pady=(0, 10))
        
        tk.Label(info_frame,
                text=f"Saved data files: {file_count}",
                font=('Inter', 11),
                fg=self.theme['text_primary'],
                bg=self.theme['bg_card']).pack(anchor='w')
        
        tk.Label(info_frame,
                text=f"Storage used: {formatted_size}",
                font=('Inter', 11),
                fg=self.theme['text_primary'],
                bg=self.theme['bg_card']).pack(anchor='w')
        
        tk.Label(info_frame,
                text=f"Location: {self.data_manager.data_dir}",
                font=('Inter', 9),
                fg=self.theme['text_muted'],
                bg=self.theme['bg_card']).pack(anchor='w', pady=(5, 0))
        
        # Clear data button
        clear_frame = tk.Frame(data_content, bg=self.theme['bg_card'])
        clear_frame.pack(fill='x', pady=(10, 0))
        
        def confirm_clear_data():
            if file_count == 0:
                messagebox.showinfo("No Data", "No saved data files to clear.")
                return
                
            result = messagebox.askyesno(
                "Clear Data", 
                f"Are you sure you want to delete all {file_count} saved data files?\n\n"
                f"This will free up {formatted_size} of storage space.\n\n"
                "This action cannot be undone.",
                icon='warning'
            )
            if result:
                if self.clear_data_folder():
                    messagebox.showinfo("Success", "All saved data files have been cleared.")
                    settings_window.destroy()
                else:
                    messagebox.showerror("Error", "Failed to clear some data files.")
        
        clear_btn = tk.Button(clear_frame,
                             text=f"Clear All Data ({file_count} files)",
                             font=('Inter', 10, 'bold'),
                             bg=self.theme['accent_red'],
                             fg='white',
                             border=0,
                             padx=20,
                             pady=8,
                             cursor='hand2',
                             activebackground='#dc2626',
                             command=confirm_clear_data,
                             state='normal' if file_count > 0 else 'disabled')
        clear_btn.pack(anchor='w')
        
        tk.Label(clear_frame,
                text="Remove all previously saved Excel file data",
                font=('Inter', 9),
                fg=self.theme['text_muted'],
                bg=self.theme['bg_card']).pack(anchor='w', padx=(0, 0), pady=(5, 0))
        
        # Close button
        button_frame = tk.Frame(main_frame, bg=self.theme['bg_dark'])
        button_frame.pack(fill='x', pady=(20, 0))
        
        close_btn = tk.Button(button_frame,
                             text="Close",
                             font=('Inter', 11, 'bold'),
                             bg=self.theme['accent_blue'],
                             fg='white',
                             border=0,
                             padx=25,
                             pady=10,
                             cursor='hand2',
                             activebackground='#2563eb',
                             command=settings_window.destroy)
        close_btn.pack(side='right')
        
        settings_window.bind('<Escape>', lambda e: settings_window.destroy())
    
    def toggle_performance_logs(self):
        self.show_performance_logs = self.perf_log_var.get()
        status = "enabled" if self.show_performance_logs else "disabled"
        print(f"Performance logging {status}")
    
    def get_data_folder_size(self):
        """Calculate the total size of the data folder in bytes"""
        try:
            data_dir = self.data_manager.data_dir
            if not os.path.exists(data_dir):
                return 0
            
            total_size = 0
            for dirpath, dirnames, filenames in os.walk(data_dir):
                for filename in filenames:
                    filepath = os.path.join(dirpath, filename)
                    try:
                        total_size += os.path.getsize(filepath)
                    except (OSError, IOError):
                        continue
            return total_size
        except Exception:
            return 0
    
    def format_file_size(self, size_bytes):
        """Convert bytes to human readable format"""
        if size_bytes == 0:
            return "0 B"
        
        size_names = ["B", "KB", "MB", "GB"]
        i = 0
        while size_bytes >= 1024 and i < len(size_names) - 1:
            size_bytes /= 1024.0
            i += 1
        
        if i == 0:
            return f"{int(size_bytes)} {size_names[i]}"
        else:
            return f"{size_bytes:.1f} {size_names[i]}"
    
    def get_data_file_count(self):
        """Count the number of saved data files"""
        try:
            data_dir = self.data_manager.data_dir
            if not os.path.exists(data_dir):
                return 0
            
            count = 0
            for filename in os.listdir(data_dir):
                if filename.endswith('.json'):
                    count += 1
            return count
        except Exception:
            return 0
    
    def clear_data_folder(self):
        """Clear all saved data files from the data folder"""
        try:
            data_dir = self.data_manager.data_dir
            if not os.path.exists(data_dir):
                return True
            
            files_deleted = 0
            for filename in os.listdir(data_dir):
                if filename.endswith('.json'):
                    filepath = os.path.join(data_dir, filename)
                    try:
                        os.remove(filepath)
                        files_deleted += 1
                    except (OSError, IOError):
                        continue
            
            print(f"Cleared {files_deleted} data files from storage")
            return True
        except Exception as e:
            print(f"Error clearing data folder: {e}")
            return False
    
    def show_day_details(self, date_key, day_trades):
        if not day_trades:
            return
        
        detail_window = tk.Toplevel(self.root)
        detail_window.title(f"Trading Details - {date_key}")
        detail_window.geometry("800x600")
        detail_window.configure(bg=self.theme['bg_dark'])
        detail_window.transient(self.root)
        detail_window.grab_set()
        
        detail_window.geometry("+{}+{}".format(
            int(self.root.winfo_screenwidth()/2 - 400),
            int(self.root.winfo_screenheight()/2 - 300)
        ))
        
        main_frame = tk.Frame(detail_window, bg=self.theme['bg_dark'])
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        header_frame = tk.Frame(main_frame, bg=self.theme['bg_dark'])
        header_frame.pack(fill='x', pady=(0, 20))
        
        date_obj = datetime.strptime(date_key, '%Y-%m-%d').date()
        formatted_date = date_obj.strftime('%A, %B %d, %Y')
        
        title_label = tk.Label(header_frame,
                              text=f"Trading Details",
                              font=('Inter', 18, 'bold'),
                              fg=self.theme['text_primary'],
                              bg=self.theme['bg_dark'])
        title_label.pack(side='left')
        
        date_label = tk.Label(header_frame,
                             text=formatted_date,
                             font=('Inter', 14),
                             fg=self.theme['text_secondary'],
                             bg=self.theme['bg_dark'])
        date_label.pack(side='left', padx=(20, 0))
        
        stats_frame = tk.Frame(main_frame, bg=self.theme['bg_card'])
        stats_frame.pack(fill='x', pady=(0, 20))
        
        stats_content = tk.Frame(stats_frame, bg=self.theme['bg_card'])
        stats_content.pack(fill='x', padx=15, pady=15)
        
        total_pnl = sum(trade['pnl'] for trade in day_trades)
        winning_trades = sum(1 for trade in day_trades if trade['pnl'] > 0)
        losing_trades = sum(1 for trade in day_trades if trade['pnl'] < 0)
        win_rate = (winning_trades / len(day_trades) * 100) if day_trades else 0
        
        total_pnl_color = self.theme['accent_green'] if total_pnl >= 0 else self.theme['accent_red']
        
        tk.Label(stats_content, text="Day Summary", font=('Inter', 12, 'bold'),
                fg=self.theme['text_primary'], bg=self.theme['bg_card']).grid(row=0, column=0, columnspan=4, pady=(0, 10))
        
        tk.Label(stats_content, text="Total P&L:", font=('Inter', 10),
                fg=self.theme['text_secondary'], bg=self.theme['bg_card']).grid(row=1, column=0, sticky='w', padx=(0, 10))
        tk.Label(stats_content, text=f"${total_pnl:,.2f}", font=('Inter', 10, 'bold'),
                fg=total_pnl_color, bg=self.theme['bg_card']).grid(row=1, column=1, sticky='w', padx=(0, 20))
        
        tk.Label(stats_content, text="Trades:", font=('Inter', 10),
                fg=self.theme['text_secondary'], bg=self.theme['bg_card']).grid(row=1, column=2, sticky='w', padx=(0, 10))
        tk.Label(stats_content, text=f"{len(day_trades)}", font=('Inter', 10, 'bold'),
                fg=self.theme['text_primary'], bg=self.theme['bg_card']).grid(row=1, column=3, sticky='w')
        
        tk.Label(stats_content, text="Winners:", font=('Inter', 10),
                fg=self.theme['text_secondary'], bg=self.theme['bg_card']).grid(row=2, column=0, sticky='w', padx=(0, 10))
        tk.Label(stats_content, text=f"{winning_trades}", font=('Inter', 10, 'bold'),
                fg=self.theme['accent_green'], bg=self.theme['bg_card']).grid(row=2, column=1, sticky='w', padx=(0, 20))
        
        tk.Label(stats_content, text="Losers:", font=('Inter', 10),
                fg=self.theme['text_secondary'], bg=self.theme['bg_card']).grid(row=2, column=2, sticky='w', padx=(0, 10))
        tk.Label(stats_content, text=f"{losing_trades}", font=('Inter', 10, 'bold'),
                fg=self.theme['accent_red'], bg=self.theme['bg_card']).grid(row=2, column=3, sticky='w')
        
        tk.Label(stats_content, text="Win Rate:", font=('Inter', 10),
                fg=self.theme['text_secondary'], bg=self.theme['bg_card']).grid(row=3, column=0, sticky='w', padx=(0, 10))
        tk.Label(stats_content, text=f"{win_rate:.1f}%", font=('Inter', 10, 'bold'),
                fg=self.theme['text_primary'], bg=self.theme['bg_card']).grid(row=3, column=1, sticky='w')
        
        trades_frame = tk.Frame(main_frame, bg=self.theme['bg_dark'])
        trades_frame.pack(fill='both', expand=True)
        
        tk.Label(trades_frame, text="Individual Trades", font=('Inter', 12, 'bold'),
                fg=self.theme['text_primary'], bg=self.theme['bg_dark']).pack(anchor='w', pady=(0, 10))
        
        trades_container = tk.Frame(trades_frame, bg=self.theme['bg_dark'])
        trades_container.pack(fill='both', expand=True)
        
        canvas = tk.Canvas(trades_container, bg=self.theme['bg_dark'], highlightthickness=0)
        scrollbar = ttk.Scrollbar(trades_container, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=self.theme['bg_dark'])
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        for i, trade in enumerate(sorted(day_trades, key=lambda x: x['pnl'], reverse=True)):
            trade_frame = tk.Frame(scrollable_frame, bg=self.theme['bg_card'], relief='flat', bd=1)
            trade_frame.pack(fill='x', pady=2, padx=2)
            
            trade_content = tk.Frame(trade_frame, bg=self.theme['bg_card'])
            trade_content.pack(fill='x', padx=10, pady=8)
            
            pnl_color = self.theme['accent_green'] if trade['pnl'] >= 0 else self.theme['accent_red']
            pnl_symbol = "+" if trade['pnl'] >= 0 else ""
            
            left_frame = tk.Frame(trade_content, bg=self.theme['bg_card'])
            left_frame.pack(side='left', fill='x', expand=True)
            
            if trade['trade_num']:
                tk.Label(left_frame, text=f"Trade #{trade['trade_num']}", font=('Inter', 10, 'bold'),
                        fg=self.theme['text_primary'], bg=self.theme['bg_card']).pack(anchor='w')
            else:
                tk.Label(left_frame, text=f"Trade #{i+1}", font=('Inter', 10, 'bold'),
                        fg=self.theme['text_primary'], bg=self.theme['bg_card']).pack(anchor='w')
            
            right_frame = tk.Frame(trade_content, bg=self.theme['bg_card'])
            right_frame.pack(side='right')
            
            tk.Label(right_frame, text=f"{pnl_symbol}${trade['pnl']:,.2f}", font=('Inter', 11, 'bold'),
                    fg=pnl_color, bg=self.theme['bg_card']).pack()
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        detail_window.bind('<Escape>', lambda e: detail_window.destroy())
    
    def create_bottom_info(self, parent):
        bottom_frame = tk.Frame(parent, bg=self.theme['bg_dark'])
        bottom_frame.pack(fill='x', pady=(20, 10))
        
        legend_frame = tk.Frame(bottom_frame, bg=self.theme['bg_dark'])
        legend_frame.pack(side='left')
        
        tk.Label(legend_frame,
                text="Legend:",
                font=('Inter', 10, 'bold'),
                fg=self.theme['text_primary'],
                bg=self.theme['bg_dark']).pack(side='left')
        
        legend_items = [
            ("● Profit", self.theme['accent_green']),
            ("● Loss", self.theme['accent_red']),
            ("● No Trading", self.theme['text_muted'])
        ]
        
        for text, color in legend_items:
            tk.Label(legend_frame,
                    text=text,
                    font=('Inter', 10),
                    fg=color,
                    bg=self.theme['bg_dark']).pack(side='left', padx=(15, 0))
    
    def load_file(self):
        file_path = filedialog.askopenfilename(
            title="Select TradingView Export",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        
        if not file_path:
            return
        
        self.current_file = file_path
            
        self.show_loading_dialog(file_path)
    
    def show_loading_dialog(self, file_path):
        self.loading_dialog = tk.Toplevel(self.root)
        self.loading_dialog.title("Processing Excel File")
        self.loading_dialog.geometry("500x350")
        self.loading_dialog.configure(bg=self.theme['bg_dark'])
        self.loading_dialog.transient(self.root)
        self.loading_dialog.grab_set()
        
        self.loading_dialog.geometry("+{}+{}".format(
            int(self.root.winfo_screenwidth()/2 - 250),
            int(self.root.winfo_screenheight()/2 - 175)
        ))
        
        self.loading_dialog.protocol("WM_DELETE_WINDOW", self.close_loading_dialog)
        self.loading_dialog.bind('<Escape>', lambda e: self.close_loading_dialog())
        self.loading_dialog.bind('<Return>', lambda e: self.close_loading_dialog())
        self.loading_dialog.attributes('-topmost', True)
        
        main_frame = tk.Frame(self.loading_dialog, bg=self.theme['bg_dark'])
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        title_label = tk.Label(main_frame,
                              text="Processing Excel File",
                              font=('Inter', 18, 'bold'),
                              fg=self.theme['accent_blue'],
                              bg=self.theme['bg_dark'])
        title_label.pack(pady=(0, 20))
        
        filename = os.path.basename(file_path)
        file_label = tk.Label(main_frame,
                             text=f"File: {filename}",
                             font=('Inter', 12),
                             fg=self.theme['text_secondary'],
                             bg=self.theme['bg_dark'])
        file_label.pack(pady=(0, 20))
        
        log_frame = tk.Frame(main_frame, bg=self.theme['bg_card'])
        log_frame.pack(fill='both', expand=True, pady=(0, 20))
        
        log_header = tk.Label(log_frame,
                             text="Processing Log:",
                             font=('Inter', 12, 'bold'),
                             fg=self.theme['text_primary'],
                             bg=self.theme['bg_card'])
        log_header.pack(anchor='w', padx=15, pady=(15, 5))
        
        self.loading_log = tk.Text(log_frame,
                                  font=('Consolas', 10),
                                  bg=self.theme['bg_dark'],
                                  fg=self.theme['text_primary'],
                                  border=0,
                                  height=10,
                                  wrap='word')
        self.loading_log.pack(fill='both', expand=True, padx=15, pady=(0, 15))
        
        self.process_file_background(file_path)
        
    def process_file_background(self, file_path):
        processing_thread = threading.Thread(
            target=self.process_file_with_logging,
            args=(file_path,),
            daemon=True
        )
        processing_thread.start()
        
    def process_file_with_logging(self, file_path):
        import time
        start_time = time.time()
        
        try:
            result = self.trade_processor.process_file(file_path, self.log_loading)
            
            end_time = time.time()
            processing_time = end_time - start_time
            
            self.trades = result['trades']
            self.daily_pnl = defaultdict(float, result['daily_pnl'])
            self.stats = result['stats']
            
            total_pnl = self.stats.get('total_pnl', 0)
            total_trades = self.stats.get('total_trades', 0)
            win_rate = self.stats.get('win_rate', 0)
            
            self.log_loading(f"Total P&L: ${total_pnl:,.2f}")
            self.log_loading(f"Total Trades: {total_trades}")
            self.log_loading(f"Win Rate: {win_rate:.1f}%")
            self.log_loading(f"Processing completed in {processing_time:.2f} seconds")
            
            self.log_loading("Preparing calendar display...")
            
            self.loading_dialog.after(0, self.finish_processing)
            
        except Exception as e:
            error_msg = str(e)
            self.log_loading(f"ERROR: {error_msg}")
            self.loading_dialog.after(0, lambda: self.show_processing_error(error_msg))
    
    def log_loading(self, message):
        def update_log():
            self.loading_log.insert('end', f"{message}\n")
            self.loading_log.see('end')
            self.loading_dialog.update()
        
        self.loading_dialog.after(0, update_log)
    
    def finish_processing(self):
        self.log_loading("Processing complete!")
        
        self.update_displays()
        
        if self.current_file and self.trades and self.stats:
            self.data_manager.save_trading_data(self.current_file, self.trades, self.daily_pnl, self.stats)
        
        filename = os.path.basename(self.current_file) if self.current_file else "Unknown"
        self.file_status.config(text=f"Loaded: {filename}")
    
    def show_processing_error(self, error_msg):
        error_btn = tk.Button(self.loading_dialog,
                             text="Close",
                             font=('Inter', 12, 'bold'),
                             bg=self.theme['accent_red'],
                             fg='white',
                             border=0,
                             padx=30,
                             pady=12,
                             cursor='hand2',
                             command=self.close_loading_dialog)
        error_btn.pack(pady=(10, 0))
        
        self.loading_dialog.after(100, lambda: messagebox.showerror("Processing Error", f"Failed to process file: {error_msg}"))
    
    def continue_to_calendar(self):
        self.close_loading_dialog()
    
    def close_loading_dialog(self):
        if hasattr(self, 'loading_dialog') and self.loading_dialog:
            self.loading_dialog.destroy()
    
    def update_displays(self):
        self.update_stats_display()
        self.cache_trades_by_date()
        self.update_calendar()
    
    def cache_trades_by_date(self):
        self.trades_by_date = {}
        for trade in self.trades:
            date_key = trade['date'].strftime('%Y-%m-%d')
            if date_key not in self.trades_by_date:
                self.trades_by_date[date_key] = []
            self.trades_by_date[date_key].append(trade)
    
    def update_stats_display(self):
        if not self.stats:
            return
        
        pnl_color = self.theme['accent_green'] if self.stats['total_pnl'] >= 0 else self.theme['accent_red']
        self.stat_cards['total_pnl'].value_label.config(
            text=f"${self.stats['total_pnl']:,.2f}",
            fg=pnl_color
        )
        
        self.stat_cards['win_rate'].value_label.config(
            text=f"{self.stats['win_rate']:.1f}%"
        )
        
        avg_color = self.theme['accent_green'] if self.stats['avg_daily'] >= 0 else self.theme['accent_red']
        self.stat_cards['avg_daily'].value_label.config(
            text=f"${self.stats['avg_daily']:.2f}",
            fg=avg_color
        )
        
        self.stat_cards['total_trades'].value_label.config(
            text=str(self.stats['total_trades'])
        )
    
    def update_calendar(self):
        start_time = time.time()
        
        if self.show_performance_logs:
            print(f"PERFORMANCE TEST - Starting calendar update...")
            print(f"Total trades: {len(self.trades)}")
        
        destroy_start = time.time()
        # OPTIMIZED: Batch widget destruction for better performance
        widgets_to_destroy = []
        for widget in self.calendar_frame.winfo_children():
            if int(widget.grid_info()['row']) > 0:
                widgets_to_destroy.append(widget)
        
        # Destroy all widgets at once
        for widget in widgets_to_destroy:
            widget.destroy()
        destroy_time = time.time() - destroy_start
        
        cal = calendar.monthcalendar(self.current_year, self.current_month)
        
        # OPTIMIZED: Pre-calculate month data to reduce lookups during cell creation
        month_data_start = time.time()
        month_pnl_data = {}
        month_trade_data = {}
        for week in cal:
            for day in week:
                if day != 0:
                    date_key = f"{self.current_year}-{self.current_month:02d}-{day:02d}"
                    month_pnl_data[day] = self.daily_pnl.get(date_key, 0)
                    month_trade_data[day] = self.trades_by_date.get(date_key, [])
        month_data_time = time.time() - month_data_start
        
        cell_creation_start = time.time()
        # OPTIMIZED: Create cells with pre-calculated data
        for week_num, week in enumerate(cal):
            for day_num, day in enumerate(week):
                if day == 0:
                    continue
                    
                row = week_num + 1
                col = day_num
                
                # Use pre-calculated data instead of lookups
                pnl = month_pnl_data[day]
                day_trades = month_trade_data[day]
                
                self.create_optimized_day_cell(day, pnl, day_trades, row, col)
        cell_creation_time = time.time() - cell_creation_start
        
        month_name = calendar.month_name[self.current_month]
        self.month_display.config(text=f"{month_name} {self.current_year}")
        
        total_time = time.time() - start_time
        
        if self.show_performance_logs:
            print(f"PERFORMANCE RESULTS:")
            print(f"  Widget destruction: {destroy_time:.3f}s")
            print(f"  Month data prep: {month_data_time:.3f}s")
            print(f"  Cell creation: {cell_creation_time:.3f}s") 
            print(f"  Total calendar update: {total_time:.3f}s")
            print(f"  Trades cache size: {len(self.trades_by_date)} days")
        
        self.performance_metrics = {
            'destroy_time': destroy_time,
            'month_data_time': month_data_time,
            'cell_creation_time': cell_creation_time,
            'total_time': total_time,
            'trade_count': len(self.trades)
        }
    
    def prev_month(self):
        if self.current_month == 1:
            self.current_month = 12
            self.current_year -= 1
        else:
            self.current_month -= 1
        self.update_calendar()
    
    def next_month(self):
        if self.current_month == 12:
            self.current_month = 1
            self.current_year += 1
        else:
            self.current_month += 1
        self.update_calendar()

# ================================ ENTRY POINT ================================

def main():
    try:
        print("Strategy Analyzer")
        print("=" * 60)
        
        root = tk.Tk()
        app = StrategyAnalyzer(root)
        
        print("Application initialized successfully")
        print("Ready to process TradingView Excel exports")
        print("=" * 60)
        
        root.mainloop()
        
    except ImportError as e:
        print(f"Missing dependency: {e}")
        print("\nThis shouldn't happen as dependencies are auto-installed.")
        print("Please ensure you have internet access and pip is available.")
        input("\nPress Enter to exit...")
        sys.exit(1)
    except Exception as e:
        print(f"Application error: {e}")
        input("\nPress any key to exit...")
        sys.exit(1)

if __name__ == "__main__":
    main()
